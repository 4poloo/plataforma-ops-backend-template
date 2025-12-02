from __future__ import annotations

import asyncio
import subprocess
import tempfile

from datetime import datetime, timezone, date, time
import csv
import io
from pathlib import Path
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple

from bson import ObjectId

from app.db.repositories import work_orders_repo, recipes_repo, products_repo, exclude_skus_repo
from app.services import wms_service
from app.models.work_orders import (
    WorkOrderContentIn,
    WorkOrderIntegrationItem,
    WorkOrderCreateIn,
    WorkOrderListFilters,
    WorkOrderEstadoUpdate,
    RecipePrintRequest,
    WorkOrderBulkImportOut,
    WorkOrderImportError,
    WorkOrderIntegrationRequest,
    WorkOrderOut,
)


def _oid_str(oid: ObjectId | None) -> str | None:
    return str(oid) if oid is not None else None


def _normalize_date(value: Any, field_name: str) -> date:
    if value is None:
        raise ValueError(f"{field_name} es requerido")

    if isinstance(value, datetime):
        return value.date()
    elif isinstance(value, date):
        return value
    elif isinstance(value, str):
        text = value.strip()
        if not text:
            raise ValueError(f"{field_name} es requerido")
        for parser in (
            lambda v: datetime.fromisoformat(v.replace("Z", "+00:00")).date(),
            lambda v: datetime.strptime(v, "%Y-%m-%d").date(),
            lambda v: datetime.strptime(v, "%d-%m-%Y").date(),
            lambda v: datetime.strptime(v, "%d/%m/%Y").date(),
            lambda v: datetime.strptime(v, "%Y/%m/%d").date(),
        ):
            try:
                return parser(text)
            except ValueError:
                continue
        else:
            raise ValueError(f"{field_name} con formato inválido: '{value}'")
    else:
        raise ValueError(f"{field_name} con formato inválido: '{value}'")
    return value


def _map_work_order(doc: Dict[str, Any]) -> Dict[str, Any]:
    contenido = doc.get("contenido", {})

    def _to_date(raw: Any) -> Optional[date]:
        if raw is None:
            return None
        if isinstance(raw, datetime):
            return raw.date()
        if isinstance(raw, date):
            return raw
        try:
            return _normalize_date(raw, "fecha")
        except ValueError:
            return None

    return {
        "_id": _oid_str(doc.get("_id")),
        "OT": doc.get("OT"),
        "contenido": {
            "SKU": contenido.get("SKU"),
            "Cantidad": contenido.get("Cantidad"),
            "Encargado": contenido.get("Encargado"),
            "linea": contenido.get("linea"),
            "fecha": _to_date(contenido.get("fecha")),
            "fecha_ini": _to_date(contenido.get("fecha_ini")),
            "fecha_fin": _to_date(contenido.get("fecha_fin")),
            "descripcion": contenido.get("descripcion"),
        },
        "estado": doc.get("estado", "CREADA"),
        "merma": float(doc.get("merma", 0) or 0),
        "cantidad_fin": float(doc.get("cantidad_fin", 0) or 0),
        "audit": doc.get("audit", {}),
    }


def _as_datetime(value: date) -> datetime:
    return datetime.combine(value, time.min, tzinfo=timezone.utc)


def _parse_float(value: Any, field: str) -> float:
    if value is None:
        raise ValueError(f"{field} es requerido")
    try:
        text = str(value).replace(",", ".").strip()
        return float(text)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{field} inválido") from exc


def _row_to_payload(row: dict) -> WorkOrderCreateIn:
    ot_raw = row.get("OT")
    if ot_raw is None:
        raise ValueError("OT es requerido")
    try:
        ot_int = int(str(ot_raw).strip())
    except (TypeError, ValueError) as exc:
        raise ValueError("OT debe ser un número entero") from exc

    contenido = WorkOrderContentIn(
        SKU=str(row.get("SKU") or "").strip(),
        Cantidad=_parse_float(row.get("Cantidad"), "Cantidad"),
        Encargado=str(row.get("Encargado") or "").strip(),
        linea=str(row.get("linea") or "").strip(),
        fecha=_normalize_date(row.get("fecha"), "fecha"),
        fecha_ini=_normalize_date(row.get("fecha_ini"), "fecha_ini"),
        fecha_fin=_normalize_date(row.get("fecha_fin"), "fecha_fin"),
        descripcion=None,
    )
    return WorkOrderCreateIn(
        OT=ot_int,
        contenido=contenido,
        estado="CREADA",
        merma=0,
        cantidad_fin=0,
    )


async def _filter_excluded_skus_from_payload(
    db,
    items: List[WorkOrderIntegrationItem],
) -> List[WorkOrderIntegrationItem]:
    if not items:
        return items

    candidate_skus = {
        str(item.CodigoMaterial).strip()
        for item in items
        if getattr(item, "CodigoMaterial", None) is not None
    }
    excluded = await exclude_skus_repo.find_matching_skus(candidate_skus, db=db)
    if not excluded:
        return items

    filtered = [
        item for item in items if str(item.CodigoMaterial).strip() not in excluded
    ]
    if not filtered:
        raise ValueError("Todos los SKUs de la OT están excluidos para integración con Invas.")
    return filtered


async def filter_wms_payload_items(
    db,
    items: List[WorkOrderIntegrationItem],
) -> List[WorkOrderIntegrationItem]:
    """Public helper to strip out materials excluded from WMS integration."""
    return await _filter_excluded_skus_from_payload(db, items)


RECIPE_EXCEL_MAP = {
    "single_cells": {
        "fechahoy": "I4",
        "numero_ot": "C10",
        "linea": "C11",
        "encargado": "C12",
        "producto": "C13",
        "fecha_ini": "I10",
    },
    "materiales_receta": {
        "rango_filas": {"inicio": 17, "fin": 35},
        "columnas": {"sku": "B", "nombre": "C", "unidad": "H", "cantidad": "I"},
    },
}

PLANTILLAS_DIR = Path(__file__).resolve().parent.parent / "plantillas"
RECIPE_TEMPLATE_PATH = PLANTILLAS_DIR / "plantillaReceta.xlsx"
RECIPE_LOGO_PATH = PLANTILLAS_DIR / "LogoEmpresa.png"

EXCEL_OT_MAP = {
    "numero_ot": "D9",
    "linea": "D10",
    "encargado": "D12",
    "producto_nombre": "D13",
    "fecha_ini": "L9",
    "fecha_fin": "L10",
    "sku": "L11",
    "cantidad": "L12",
}

OT_TEMPLATE_PATH = PLANTILLAS_DIR / "plantilla_OT.xlsx"
OT_LOGO_PATH = PLANTILLAS_DIR / "LogoEmpresa.png"


async def create_work_order(db, payload) -> Dict[str, Any]:
    existing = await work_orders_repo.find_by_ot(payload.OT, db=db)
    if existing:
        raise ValueError(f"La OT {payload.OT} ya existe")

    contenido = payload.contenido
    fecha = _normalize_date(contenido.fecha, "fecha")
    fecha_ini = _normalize_date(contenido.fecha_ini, "fecha_ini")
    fecha_fin = _normalize_date(contenido.fecha_fin, "fecha_fin")

    if fecha_ini > fecha_fin:
        raise ValueError("fecha_ini no puede ser mayor a fecha_fin")

    now = datetime.now(timezone.utc)
    descripcion = (
        contenido.descripcion.strip()
        if isinstance(contenido.descripcion, str) and contenido.descripcion.strip()
        else None
    )

    work_order_doc: Dict[str, Any] = {
        "OT": int(payload.OT),
        "contenido": {
            "SKU": contenido.SKU,
            "Cantidad": float(contenido.Cantidad),
            "Encargado": contenido.Encargado,
            "linea": contenido.linea,
            "fecha": _as_datetime(fecha),
            "fecha_ini": _as_datetime(fecha_ini),
            "fecha_fin": _as_datetime(fecha_fin),
            "descripcion": descripcion,
        },
        "estado": payload.estado,
        "merma": float(payload.merma),
        "cantidad_fin": float(payload.cantidad_fin),
        "audit": {"createdAt": now, "updatedAt": now},
    }

    saved = await work_orders_repo.insert_work_order(work_order_doc, db=db)
    return _map_work_order(saved)


async def list_work_orders(
    db,
    *,
    limit: int = 50,
    skip: int = 0,
    filters: WorkOrderListFilters | None = None,
) -> list[Dict[str, Any]]:
    filtro_db: Dict[str, Any] = {}
    if filters and filters.estado:
        filtro_db["estado"] = filters.estado

    docs = await work_orders_repo.list_work_orders(
        db=db,
        limit=limit,
        skip=skip,
        filtro=filtro_db or None,
        sort=[("OT", 1)],
    )
    return [_map_work_order(d) for d in docs]


async def get_next_ot(db) -> int:
    """Obtiene la siguiente OT disponible tomando la mayor OT almacenada."""
    last = await work_orders_repo.find_last_ot(db=db)
    if not last or last.get("OT") is None:
        return 1
    try:
        return int(last["OT"]) + 1
    except (TypeError, ValueError):
        return 1


async def get_last_created_ot(db) -> int:
    """Obtiene el número de la última OT creada."""
    last = await work_orders_repo.find_last_created(db=db)
    if not last or last.get("OT") is None:
        raise ValueError("No hay órdenes de trabajo registradas")
    try:
        return int(last["OT"])
    except (TypeError, ValueError) as exc:
        raise ValueError("El número de OT almacenado es inválido") from exc


async def get_work_order_by_ot(db, ot: int | str) -> Dict[str, Any]:
    try:
        ot_int = int(ot)
    except (TypeError, ValueError) as exc:
        raise ValueError("El número de OT debe ser un entero") from exc

    doc = await work_orders_repo.find_by_ot(ot_int, db=db)
    if not doc:
        raise ValueError("OT no encontrada")
    return _map_work_order(doc)


async def import_work_orders_from_csv(db, csv_text: str) -> WorkOrderBulkImportOut:
    reader = csv.DictReader(io.StringIO(csv_text))
    required = {"OT", "SKU", "Cantidad", "Encargado", "linea", "fecha", "fecha_ini", "fecha_fin"}
    if reader.fieldnames is None or not required.issubset({(h or "").strip() for h in reader.fieldnames}):
        raise ValueError("CSV inválido: faltan cabeceras requeridas")

    errors: list[WorkOrderImportError] = []
    pending_to_save: list[tuple[int, WorkOrderCreateIn]] = []
    integration_items: list[WorkOrderIntegrationItem] = []

    for idx, row in enumerate(reader, start=2):  # empieza en 2 por el header
        try:
            payload = _row_to_payload(row)
            # Construye items para enviar a WMS (sin guardar aún en BD)
            items = await build_wms_integration_items(
                db,
                ot=payload.OT,
                contenido=payload.contenido,
            )
            if not items:
                raise ValueError("Receta sin componentes válidos para enviar a Invas")
            integration_items.extend(items)
            pending_to_save.append((idx, payload))
        except Exception as exc:  # noqa: BLE001
            ot_val = None
            try:
                ot_val = int(str(row.get("OT") or "").strip())
            except Exception:
                ot_val = None
            errors.append(WorkOrderImportError(row=idx, ot=ot_val, error=str(exc)))

    wms_response = None
    wms_error = None
    wms_error_detail = None
    created: list[Dict[str, Any]] = []

    # Primero intentamos enviar a Invas; si falla, no persistimos las OT
    if integration_items:
        try:
            wms_response = await wms_service.send_work_orders(
                WorkOrderIntegrationRequest(
                    source="portal",
                    payload=integration_items,
                )
            )
        except wms_service.WMSIntegrationError as exc:
            wms_error = str(exc)
            wms_error_detail = exc.body
        except Exception as exc:  # noqa: BLE001
            wms_error = str(exc)

    if wms_response and not wms_error:
        # Solo guardamos en BD si el envío a Invas fue exitoso
        for idx, payload in pending_to_save:
            try:
                created_doc = await create_work_order(db, payload)
                created.append(created_doc)
            except Exception as exc:  # noqa: BLE001
                errors.append(WorkOrderImportError(row=idx, ot=payload.OT, error=str(exc)))

    created_out = [WorkOrderOut(**doc) for doc in created]
    return WorkOrderBulkImportOut(
        created=created_out,
        errors=errors,
        wms_response=wms_response,
        wms_error=wms_error,
        wms_error_detail=wms_error_detail,
    )


async def update_work_order_estado(
    db,
    ot: int | str,
    payload: WorkOrderEstadoUpdate,
) -> Dict[str, Any]:
    try:
        ot_int = int(ot)
    except (TypeError, ValueError) as exc:
        raise ValueError("El número de OT debe ser un entero") from exc

    updated = await work_orders_repo.update_estado_by_ot(
        ot_int,
        payload.estado,
        db=db,
    )
    if not updated:
        raise ValueError("OT no encontrada")
    return _map_work_order(updated)


def _format_ddmmyyyy(value: date) -> str:
    return value.strftime("%d/%m/%Y")


def _format_iso_date(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        value = value.date()
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    return str(value)


async def generate_work_order_excel(db, ot: int | str) -> Tuple[str, BytesIO]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl no está instalado. Instálalo para generar la plantilla de OT."
        ) from exc

    try:
        from openpyxl.drawing.image import Image
    except ModuleNotFoundError:
        Image = None

    if not OT_TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"No se encontró la plantilla en {OT_TEMPLATE_PATH}")
    if Image is not None and not OT_LOGO_PATH.exists():
        raise FileNotFoundError(f"No se encontró el logo en {OT_LOGO_PATH}")

    order = await get_work_order_by_ot(db, ot)
    contenido = order.get("contenido", {})

    producto = await products_repo.find_by_sku(
        contenido.get("SKU"),
        db=db,
        projection={"nombre": 1},
    )
    producto_nombre = (producto or {}).get("nombre") or contenido.get("SKU", "")

    value_map = {
        "numero_ot": str(order.get("OT", "")),
        "linea": contenido.get("linea", ""),
        "encargado": contenido.get("Encargado", ""),
        "producto_nombre": producto_nombre,
        "fecha_ini": _format_iso_date(contenido.get("fecha_ini")),
        "fecha_fin": _format_iso_date(contenido.get("fecha_fin")),
        "sku": contenido.get("SKU", ""),
        "cantidad": float(contenido.get("Cantidad", 0) or 0),
    }

    workbook = load_workbook(filename=OT_TEMPLATE_PATH)
    sheet = workbook.active

    if Image is not None:
        try:
            logo_img = Image(str(OT_LOGO_PATH))
            col_b = sheet.column_dimensions.get("B")
            col_c = sheet.column_dimensions.get("C")
            width_b = (col_b.width if col_b and col_b.width else 8.43) * 7
            width_c = (col_c.width if col_c and col_c.width else 8.43) * 7
            total_width = width_b + width_c
            row2 = sheet.row_dimensions.get(2)
            height_row = (row2.height if row2 and row2.height else 15) * 1.35
            # Celda B2:C5 -> 4 filas
            row3 = sheet.row_dimensions.get(3)
            row4 = sheet.row_dimensions.get(4)
            row5 = sheet.row_dimensions.get(5)
            height_row += (row3.height if row3 and row3.height else 15) * 1.35
            height_row += (row4.height if row4 and row4.height else 15) * 1.35
            height_row += (row5.height if row5 and row5.height else 15) * 1.35
            logo_img.width = total_width
            logo_img.height = height_row
            sheet.add_image(logo_img, "B2")
        except Exception as exc:
            raise RuntimeError(f"No se pudo insertar el logo en la plantilla de OT: {exc}") from exc
    for key, cell in EXCEL_OT_MAP.items():
        sheet[cell].value = value_map.get(key, "")

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    filename = f"OT_{value_map['numero_ot']}.xlsx"
    return filename, buffer


def _convert_excel_bytes_to_pdf(excel_bytes: bytes, source_filename: str) -> Tuple[str, bytes]:
    """
    Convierte un archivo Excel (bytes) a PDF utilizando LibreOffice en modo headless.
    """
    with tempfile.TemporaryDirectory() as tmp_dir:
        tmp_dir_path = Path(tmp_dir)
        xlsx_path = tmp_dir_path / source_filename
        xlsx_path.write_bytes(excel_bytes)

        output_dir = tmp_dir_path
        conversion_errors: list[str] = []
        for executable in ("libreoffice", "soffice"):
            try:
                subprocess.run(
                    [
                        executable,
                        "--headless",
                        "--convert-to",
                        "pdf",
                        "--outdir",
                        str(output_dir),
                        str(xlsx_path),
                    ],
                    check=True,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.PIPE,
                )
                break
            except FileNotFoundError:
                conversion_errors.append(
                    f"No se encontró el ejecutable '{executable}' en el sistema."
                )
                continue
            except subprocess.CalledProcessError as exc:
                stderr = exc.stderr.decode("utf-8", errors="ignore") if exc.stderr else ""
                conversion_errors.append(
                    f"Error de LibreOffice ({executable}): {stderr.strip() or exc}"
                )
                continue
        else:
            raise RuntimeError(
                "No fue posible convertir el Excel a PDF. "
                + (" ".join(conversion_errors) or "Verifica la instalación de LibreOffice.")
            )

        pdf_path = output_dir / Path(source_filename).with_suffix(".pdf").name
        if not pdf_path.exists():
            raise RuntimeError(
                "La conversión a PDF no generó ningún archivo. Revisa la plantilla de OT."
            )

        return pdf_path.name, pdf_path.read_bytes()


async def generate_work_order_pdf(db, ot: int | str) -> Tuple[str, BytesIO]:
    filename_xlsx, buffer = await generate_work_order_excel(db, ot)
    buffer.seek(0)
    excel_bytes = buffer.read()

    pdf_filename, pdf_bytes = await asyncio.to_thread(
        _convert_excel_bytes_to_pdf,
        excel_bytes,
        filename_xlsx,
    )

    pdf_buffer = BytesIO(pdf_bytes)
    pdf_buffer.seek(0)
    return pdf_filename, pdf_buffer


async def generate_recipe_excel(db, payload: RecipePrintRequest) -> Tuple[str, BytesIO]:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise RuntimeError("openpyxl no está instalado. Instálalo para generar la plantilla de receta.") from exc

    try:
        from openpyxl.drawing.image import Image
    except ModuleNotFoundError:
        Image = None

    if not RECIPE_TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"No se encontró la plantilla en {RECIPE_TEMPLATE_PATH}")
    if Image is not None and not RECIPE_LOGO_PATH.exists():
        raise FileNotFoundError(f"No se encontró el logo en {RECIPE_LOGO_PATH}")

    sku_pt = payload.skuPT
    pt_doc = await products_repo.find_by_sku(sku_pt, db=db)
    if not pt_doc:
        raise ValueError(f"PT no encontrado para sku '{sku_pt}'")

    receta_doc = await recipes_repo.find_by_pt_id(pt_doc["_id"], db=db)
    if not receta_doc or not receta_doc.get("versiones"):
        raise ValueError(f"Receta no encontrada para sku '{sku_pt}'")

    versiones = receta_doc.get("versiones", [])
    version_doc = None
    vigente = receta_doc.get("vigenteVersion")
    if vigente is not None:
        version_doc = next((v for v in versiones if int(v.get("version")) == int(vigente)), None)
    if version_doc is None and versiones:
        version_doc = sorted(versiones, key=lambda v: int(v.get("version", 0)), reverse=True)[0]
    if version_doc is None:
        raise ValueError(f"Receta sin versiones válidas para sku '{sku_pt}'")

    componentes = version_doc.get("componentes") or []
    if not componentes:
        raise ValueError(f"Receta sin componentes para sku '{sku_pt}'")

    base_qty = float(version_doc.get("base_qty") or 1.0)
    if base_qty <= 0:
        base_qty = 1.0
    qty_target = base_qty if payload.cantidad is None else float(payload.cantidad)
    factor = qty_target / base_qty

    component_ids: List[ObjectId] = []
    for comp in componentes:
        pid = comp.get("productId")
        if not pid:
            continue
        if not isinstance(pid, ObjectId):
            try:
                pid = ObjectId(pid)
                comp["productId"] = pid
            except Exception as exc:
                raise ValueError(f"productId inválido en componente: {pid}") from exc
        component_ids.append(pid)

    productos_componentes = await recipes_repo.find_products_by_ids(
        component_ids,
        db=db,
        projection={"sku": 1, "nombre": 1},
    )
    mapa_productos = {str(prod["_id"]): prod for prod in productos_componentes}

    filas = []
    for comp in componentes:
        pid = comp.get("productId")
        if not pid:
            continue
        prod_info = mapa_productos.get(str(pid))
        if not prod_info:
            raise ValueError(f"Producto no encontrado para componente {pid}")
        qty_base = float(comp.get("cantidadPorBase") or 0.0)
        merma_pct = float(comp.get("merma_pct") or 0.0)
        qty_material = qty_base * factor
        if merma_pct:
            qty_material *= (1.0 + merma_pct / 100.0)
        if qty_material < 0:
            raise ValueError(f"Cantidad calculada inválida para componente {prod_info.get('sku')}")
        filas.append(
            {
                "sku": prod_info.get("sku") or "",
                "nombre": prod_info.get("nombre") or "",
                "unidad": comp.get("unidad") or "",
                "cantidad": round(qty_material, 6),
            }
        )

    rango = RECIPE_EXCEL_MAP["materiales_receta"]["rango_filas"]
    inicio, fin = rango["inicio"], rango["fin"]
    max_filas = fin - inicio + 1
    if len(filas) > max_filas:
        raise ValueError(f"La receta tiene más componentes ({len(filas)}) que filas disponibles ({max_filas}).")

    workbook = load_workbook(filename=RECIPE_TEMPLATE_PATH)
    sheet = workbook.active

    if Image is not None:
        try:
            logo_img = Image(str(RECIPE_LOGO_PATH))
            col_width = sheet.column_dimensions.get("B")
            col_pixels = (col_width.width if col_width and col_width.width else 8.43) * 7
            row_dim = sheet.row_dimensions.get(2)
            row_pixels = (row_dim.height if row_dim and row_dim.height else 15) * 1.35
            logo_img.width = col_pixels
            logo_img.height = row_pixels
            sheet.add_image(logo_img, "B2")
        except Exception as exc:
            raise RuntimeError(f"No se pudo insertar el logo en la plantilla: {exc}") from exc

    hoy = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    producto_nombre = pt_doc.get("nombre") or sku_pt

    single_values = {
        "fechahoy": hoy,
        "numero_ot": payload.numeroOT or "",
        # 'linea' se deja intacta en la plantilla
        "encargado": payload.encargado or "",
        "producto": producto_nombre,
        "fecha_ini": _format_iso_date(payload.fecha_ini or date.today()),
    }

    for key, cell in RECIPE_EXCEL_MAP["single_cells"].items():
        if key not in single_values:
            continue
        sheet[cell].value = single_values[key]

    columnas = RECIPE_EXCEL_MAP["materiales_receta"]["columnas"]
    current_row = inicio
    for fila in filas:
        sheet[f"{columnas['sku']}{current_row}"].value = fila["sku"]
        sheet[f"{columnas['nombre']}{current_row}"].value = fila["nombre"]
        sheet[f"{columnas['unidad']}{current_row}"].value = fila["unidad"]
        sheet[f"{columnas['cantidad']}{current_row}"].value = fila["cantidad"]
        current_row += 1

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    ot_part = payload.numeroOT or "sin_OT"
    filename = f"Receta_{sku_pt}_OT_{ot_part}.xlsx"
    return filename, buffer


async def generate_recipe_pdf(db, payload: RecipePrintRequest) -> Tuple[str, BytesIO]:
    filename_xlsx, buffer = await generate_recipe_excel(db, payload)
    buffer.seek(0)
    excel_bytes = buffer.read()

    pdf_filename, pdf_bytes = await asyncio.to_thread(
        _convert_excel_bytes_to_pdf,
        excel_bytes,
        filename_xlsx,
    )

    pdf_buffer = BytesIO(pdf_bytes)
    pdf_buffer.seek(0)
    return pdf_filename, pdf_buffer


async def build_wms_integration_items(
    db,
    *,
    ot: int,
    contenido: WorkOrderContentIn,
) -> List[WorkOrderIntegrationItem]:
    sku_pt = contenido.SKU
    pt_doc = await recipes_repo.get_pt_by_sku(sku_pt, db)
    if not pt_doc:
        raise ValueError(f"PT no encontrado para sku '{sku_pt}'")

    receta_doc = await recipes_repo.find_by_pt_id(pt_doc["_id"], db=db)
    if not receta_doc or not receta_doc.get("versiones"):
        raise ValueError(f"Receta no encontrada para sku '{sku_pt}'")

    versiones = receta_doc.get("versiones", [])
    version_doc = None
    vigente = receta_doc.get("vigenteVersion")
    if vigente is not None:
        version_doc = next(
            (v for v in versiones if int(v.get("version")) == int(vigente)),
            None,
        )
    if version_doc is None:
        versiones_sorted = sorted(
            versiones,
            key=lambda v: int(v.get("version", 0)),
            reverse=True,
        )
        version_doc = versiones_sorted[0] if versiones_sorted else None
    if version_doc is None:
        raise ValueError(f"Receta sin versiones válidas para sku '{sku_pt}'")

    componentes = version_doc.get("componentes") or []
    if not componentes:
        raise ValueError(f"Receta sin componentes para sku '{sku_pt}'")

    component_ids: List[ObjectId] = []
    for comp in componentes:
        pid = comp.get("productId")
        if not pid:
            continue
        if not isinstance(pid, ObjectId):
            try:
                pid = ObjectId(pid)
                comp["productId"] = pid
            except Exception as exc:
                raise ValueError(f"productId inválido en componente: {pid}") from exc
        component_ids.append(pid)
    if not component_ids:
        raise ValueError("Componentes sin productId asociado en la receta.")

    productos_componentes = await recipes_repo.find_products_by_ids(
        component_ids,
        db=db,
        projection={"sku": 1, "nombre": 1},
    )
    mapa_componentes = {str(prod["_id"]): prod for prod in productos_componentes}

    missing = [pid for pid in component_ids if str(pid) not in mapa_componentes]
    if missing:
        raise ValueError(f"Productos no encontrados para los componentes: {', '.join(map(str, missing))}")

    pt_nombre = pt_doc.get("nombre") or ""
    if not pt_nombre:
        # fallback al repositorio de productos si el campo no viene
        prod_completo = await products_repo.find_by_sku(sku_pt, db=db, projection={"nombre": 1})
        pt_nombre = (prod_completo or {}).get("nombre") or ""

    base_qty = float(version_doc.get("base_qty") or 1.0)
    if base_qty <= 0:
        base_qty = 1.0
    qty_ot = float(contenido.Cantidad)
    factor = qty_ot / base_qty

    fecha_inicio = contenido.fecha_ini or contenido.fecha
    if fecha_inicio is None:
        fecha_inicio = contenido.fecha_fin
    if fecha_inicio is None:
        fecha_inicio = date.today()
    fec_ini = _format_ddmmyyyy(fecha_inicio)

    glosa = contenido.Encargado or contenido.linea or (pt_nombre or sku_pt)

    items: List[WorkOrderIntegrationItem] = []
    for comp in componentes:
        prod_id = comp.get("productId")
        prod_info = mapa_componentes.get(str(prod_id))
        if not prod_info:
            raise ValueError(f"Producto de componente no encontrado (productId={prod_id})")

        qty_base = float(comp.get("cantidadPorBase") or 0.0)
        merma_pct = float(comp.get("merma_pct") or 0.0)

        qty_material = qty_base * factor
        if merma_pct:
            qty_material *= (1.0 + merma_pct / 100.0)
        if qty_material <= 0:
            # Si la receta trae cantidad 0, simplemente no se envía ese componente a Invas.
            continue

        descripcion_material = prod_info.get("nombre") or prod_info.get("sku") or ""
        codigo_material = prod_info.get("sku")
        if not codigo_material:
            raise ValueError(f"Producto de componente sin SKU (productId={prod_id})")

        items.append(
            WorkOrderIntegrationItem(
                FecIniOrden=fec_ini,
                GlosaOrden=glosa,
                Orden=str(ot),
                CodigoProducto=sku_pt,
                DescripcionProducto=pt_nombre or sku_pt,
                CantidadAFabricar=qty_ot,
                CodigoMaterial=codigo_material,
                DescripcionMaterial=descripcion_material,
                CantidadMaterial=round(qty_material, 6),
            )
        )

    return await _filter_excluded_skus_from_payload(db, items)
